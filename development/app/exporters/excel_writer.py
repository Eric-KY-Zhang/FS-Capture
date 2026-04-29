"""Write the FS Capture workbook.

The report sheets follow the older VBA workbook's analyst-friendly layout:
each row is one company-period, columns A-C are stock code / name / report date,
and financial statement lines expand horizontally from column D.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import json
from pathlib import Path
import re
from typing import Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.job import Job, TaskResult, TaskStatus
from app.core.models import Exchange, Period, StatementType, Ticker


_PRIMARY_FONT = Font(name="Microsoft YaHei", size=10)
_HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="FF4338CA")
_SECTION_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FF0F172A")
_SECTION_FILL = PatternFill("solid", fgColor="FFE0E7FF")
_THIN = Side(style="thin", color="FFE6E8EC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NUMFMT = "#,##0.00;(#,##0.00);-"
_PERCENT_INDICATORS = {"ROE", "毛利率", "资产负债率"}

_SCHEMA_CACHE: dict | None = None
_LEADING_SECTION_RE = re.compile(r"^[一二三四五六七八九十]+[、.．]\s*")

_STATUS_LABEL = {
    TaskStatus.PENDING: "待处理",
    TaskStatus.RESOLVING: "识别公司",
    TaskStatus.DOWNLOADING: "下载报告",
    TaskStatus.SCRAPING: "抓取财务数据",
    TaskStatus.DONE: "成功",
    TaskStatus.FAILED: "失败",
    TaskStatus.CANCELLED: "已取消",
}

_LINE_LABELS = {
    "TOTAL_ASSETS": "资产总计",
    "TOTAL_LIABILITIES": "负债合计",
    "TOTAL_EQUITY": "股东权益合计",
    "TOTAL_CURRENT_ASSETS": "流动资产合计",
    "TOTAL_CURRENT_LIAB": "流动负债合计",
    "MONETARYFUNDS": "货币资金",
    "ACCOUNTS_RECE": "应收账款",
    "INVENTORY": "存货",
    "FIXED_ASSET": "固定资产",
    "TOTAL_OPERATE_INCOME": "营业收入",
    "OPERATE_INCOME": "营业收入",
    "OPERATE_COST": "营业成本",
    "OPERATE_PROFIT": "营业利润",
    "TOTAL_PROFIT": "利润总额",
    "NETPROFIT": "净利润",
    "PARENT_NETPROFIT": "归母净利润",
    "BASIC_EPS": "基本每股收益",
    "NETCASH_OPERATE": "经营活动现金流量净额",
    "NETCASH_INVEST": "投资活动现金流量净额",
    "NETCASH_FINANCE": "筹资活动现金流量净额",
    "SALES_SERVICES": "销售商品、提供劳务收到的现金",
}

_LOOKUP_ALIASES = {
    "营业收入": ["营业额", "营运收入", "主营业务收入", "TOTAL_OPERATE_INCOME", "OPERATE_INCOME", "Revenues"],
    "营业成本": ["营运支出", "主营业务成本", "OPERATE_COST", "CostOfRevenue", "CostOfGoodsAndServicesSold"],
    "营业利润": ["经营溢利", "OPERATE_PROFIT", "OperatingIncomeLoss"],
    "利润总额": ["除税前溢利", "TOTAL_PROFIT"],
    "净利润": ["除税后溢利", "股东应占溢利", "PARENT_NETPROFIT", "NETPROFIT", "NetIncomeLoss"],
    "归属于母公司所有者的净利润": ["股东应占溢利", "PARENT_NETPROFIT"],
    "基本每股收益": ["每股基本盈利", "BASIC_EPS", "EarningsPerShareBasic"],
    "稀释每股收益": ["每股摊薄盈利", "EarningsPerShareDiluted"],
    "资产总计": ["总资产", "资产总额", "TOTAL_ASSETS", "Assets"],
    "负债合计": ["总负债", "TOTAL_LIABILITIES", "Liabilities"],
    "股东权益合计": ["股东权益", "总权益", "TOTAL_EQUITY", "StockholdersEquity"],
    "所有者权益(或股东权益)合计": ["总权益", "股东权益", "StockholdersEquity"],
    "货币资金": ["现金及等价物", "现金及现金等价物", "MONETARYFUNDS", "CashAndCashEquivalentsAtCarryingValue"],
    "应收账款": ["应收帐款", "ACCOUNTS_RECE", "AccountsReceivableNetCurrent"],
    "流动资产合计": ["TOTAL_CURRENT_ASSETS", "AssetsCurrent"],
    "流动负债合计": ["TOTAL_CURRENT_LIAB", "LiabilitiesCurrent"],
    "长期借款": ["长期贷款", "LongTermDebtNoncurrent"],
    "短期借款": ["短期贷款", "ShortTermBorrowings", "LongTermDebtCurrent"],
    "经营活动现金流量净额": ["经营业务现金净额", "经营活动产生的现金流量净额", "NETCASH_OPERATE", "NetCashProvidedByUsedInOperatingActivities"],
    "投资活动现金流量净额": ["投资业务现金净额", "NETCASH_INVEST", "NetCashProvidedByUsedInInvestingActivities"],
    "筹资活动现金流量净额": ["融资业务现金净额", "NETCASH_FINANCE", "NetCashProvidedByUsedInFinancingActivities"],
    "现金净增加额": ["现金净额", "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect"],
    "销售商品、提供劳务收到的现金": ["SALES_SERVICES"],
    "经营活动现金流入小计": ["TOTAL_OPERATE_INFLOW"],
    "购买商品、接受劳务支付的现金": ["BUY_SERVICES"],
    "经营活动现金流出小计": ["TOTAL_OPERATE_OUTFLOW"],
    "收回投资所收到的现金": ["WITHDRAW_INVEST"],
    "取得投资收益所收到的现金": ["RECEIVE_INVEST_INCOME"],
}


def _load_schema() -> dict:
    global _SCHEMA_CACHE
    if _SCHEMA_CACHE is not None:
        return _SCHEMA_CACHE
    schema_path = Path(__file__).with_name("templates") / "sina_v22_schema.json"
    try:
        _SCHEMA_CACHE = json.loads(schema_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning(f"failed to load workbook schema {schema_path}: {exc}")
        _SCHEMA_CACHE = {}
    return _SCHEMA_CACHE


def _schema_columns(sheet_title: str) -> list[str]:
    labels = (_load_schema().get(sheet_title) or {}).get("labels") or []
    out: list[str] = []
    seen: set[str] = set()
    for label in labels[3:]:
        if isinstance(label, str) and label not in seen:
            out.append(label)
            seen.add(label)
    return out


def _apply_header(cell) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDER


def _apply_section(cell) -> None:
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDER


def _apply_value(cell, *, numeric: bool = True) -> None:
    cell.font = _PRIMARY_FONT
    cell.alignment = Alignment(horizontal="right" if numeric else "left", vertical="center")
    cell.border = _BORDER
    if numeric:
        cell.number_format = _NUMFMT


def _period_date(p: Period) -> str:
    md = {
        "annual": "12-31",
        "q1": "03-31",
        "q2": "06-30",
        "q3": "09-30",
    }[p.type.value]
    return f"{p.year}-{md}"


def _line_label(name: str, used: set[str]) -> str:
    label = _LINE_LABELS.get(name, name)
    if label in used and label != name:
        label = f"{label} ({name})"
    used.add(label)
    return label


def _strip_unit_suffix(name: str) -> str:
    return re.sub(r"[（(][^）)]*[）)]$", "", name).strip()


def _norm_label(name: str) -> str:
    name = _strip_unit_suffix(_LEADING_SECTION_RE.sub("", str(name)).strip())
    replacements = {
        "帳": "账",
        "賬": "账",
        "其它": "其他",
        "總": "总",
        "淨": "净",
        "應": "应",
        "營": "营",
    }
    for old, new in replacements.items():
        name = name.replace(old, new)
    return re.sub(r"[\s_（）()：:、，,；;\-]+", "", name)


def _lookup_line(lines: dict, name: str) -> Optional[float]:
    candidates = [name, _strip_unit_suffix(name)]
    candidates.extend(_LOOKUP_ALIASES.get(name, []))
    candidates.extend(_LOOKUP_ALIASES.get(_strip_unit_suffix(name), []))
    candidates.extend(raw for raw, label in _LINE_LABELS.items() if label in candidates)
    for key in candidates:
        if key in lines and lines[key] is not None:
            return lines[key]

    normalized = {_norm_label(k): v for k, v in lines.items() if v is not None}
    for key in candidates:
        value = normalized.get(_norm_label(key))
        if value is not None:
            return value
    return None


def _value(lines: dict, *keys: str) -> Optional[float]:
    for key in keys:
        value = _lookup_line(lines, key)
        if value is not None:
            return value
    return None


def _safe_div(a, b) -> Optional[float]:
    try:
        if a is None or b in (None, 0):
            return None
        return float(a) / float(b)
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def _scale_billion(v: Optional[float]) -> Optional[float]:
    try:
        return None if v is None else float(v) / 1e8
    except (TypeError, ValueError):
        return None


def _scale_percent(v: Optional[float]) -> Optional[float]:
    try:
        return None if v is None else float(v) * 100
    except (TypeError, ValueError):
        return None


def _gross_margin(lines_is: dict) -> Optional[float]:
    rev = _value(
        lines_is,
        "营业收入", "TOTAL_OPERATE_INCOME", "OPERATE_INCOME",
        "Revenues", "RevenueFromContractWithCustomerExcludingAssessedTax",
    )
    cost = _value(lines_is, "营业成本", "OPERATE_COST", "CostOfRevenue", "CostOfGoodsAndServicesSold")
    if rev is None or cost is None or rev == 0:
        return None
    return (float(rev) - float(cost)) / float(rev)


def _columns_in_order(results: list[TaskResult]) -> list[tuple[Ticker, Period]]:
    seen: dict[tuple[str, str, int, str], tuple[Ticker, Period]] = {}
    for r in results:
        key = (r.ticker.exchange.value, r.ticker.code, r.period.year, r.period.type.value)
        seen[key] = (r.ticker, r.period)
    return [seen[k] for k in sorted(seen.keys())]


def _build_sample_pool(ws, results: list[TaskResult]) -> None:
    ws.title = "样本池"
    headers = ["序号", "交易所", "代码", "公司名称", "外部ID", "期间数", "成功任务数", "失败任务数", "下载报告数", "财务报表数", "错误摘要"]
    for i, h in enumerate(headers, 1):
        _apply_header(ws.cell(row=1, column=i, value=h))

    by_ticker: dict[tuple[Exchange, str], list[TaskResult]] = defaultdict(list)
    for r in results:
        by_ticker[(r.ticker.exchange, r.ticker.code)].append(r)

    row = 2
    for idx, ((exch, code), tasks) in enumerate(sorted(by_ticker.items()), 1):
        ticker = tasks[0].ticker
        values = [
            idx,
            exch.display_name,
            code,
            ticker.name or "",
            ticker.external_id or "",
            len(tasks),
            sum(t.status is TaskStatus.DONE for t in tasks),
            sum(t.status is TaskStatus.FAILED for t in tasks),
            sum(len(t.reports) for t in tasks),
            sum(len(t.statements) for t in tasks),
            "；".join(t.error or "" for t in tasks if t.error),
        ]
        for col, val in enumerate(values, 1):
            _apply_value(ws.cell(row=row, column=col, value=val), numeric=isinstance(val, int))
        row += 1

    widths = [6, 8, 12, 32, 18, 10, 10, 10, 10, 10, 56]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _build_company_info(ws, results: list[TaskResult]) -> None:
    ws.title = "上市公司基本资料"
    headers = ["交易所", "代码", "公司名称", "币种", "上市日期", "行业", "外部ID", "任务状态"]
    for i, h in enumerate(headers, 1):
        _apply_header(ws.cell(row=1, column=i, value=h))

    seen: set[tuple[Exchange, str]] = set()
    row = 2
    for r in results:
        key = (r.ticker.exchange, r.ticker.code)
        if key in seen:
            continue
        seen.add(key)
        company = r.company
        currency = company.currency if company else (r.statements[0].currency if r.statements else "")
        values = [
            r.ticker.exchange.display_name,
            r.ticker.code,
            r.ticker.name or "",
            currency,
            company.listing_date if company else "",
            company.industry if company else "",
            r.ticker.external_id or "",
            _STATUS_LABEL.get(r.status, r.status.value),
        ]
        for col, val in enumerate(values, 1):
            _apply_value(ws.cell(row=row, column=col, value=val), numeric=False)
        row += 1

    widths = [8, 12, 32, 8, 14, 24, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _build_statement_sheet(ws, title: str, statement_type: StatementType, results: list[TaskResult]) -> None:
    ws.title = title
    source_line_names: list[str] = []
    seen_lines: set[str] = set()
    lines_by_task: dict[int, dict[str, Optional[float]]] = {}

    for r in results:
        for statement in r.statements:
            if statement.statement_type is not statement_type:
                continue
            lines_by_task[id(r)] = statement.lines
            for name in statement.lines:
                if name not in seen_lines:
                    seen_lines.add(name)
                    source_line_names.append(name)

    schema_names = _schema_columns(title)
    line_names = schema_names + [
        name for name in source_line_names
        if name not in schema_names and _line_label(name, set()) not in schema_names
    ]

    if not line_names:
        ws.cell(row=1, column=1, value="暂无财务数据")
        ws.cell(row=2, column=1, value="请查看“样本池”的任务状态和错误摘要，确认该期间是否抓到了财务报表。")
        ws.column_dimensions["A"].width = 90
        return

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    _apply_section(ws.cell(row=1, column=1, value="基础信息"))
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=3 + len(line_names))
    _apply_section(ws.cell(row=1, column=4, value="报表项目"))

    used_labels: set[str] = set()
    headers = ["股票代码", "名称", "报告日期"] + [_line_label(name, used_labels) for name in line_names]
    for col, h in enumerate(headers, 1):
        _apply_header(ws.cell(row=2, column=col, value=h))

    row = 3
    for r in results:
        lines = lines_by_task.get(id(r))
        if not lines:
            continue
        for col, val in enumerate([r.ticker.code, r.ticker.name or "", _period_date(r.period)], 1):
            _apply_value(ws.cell(row=row, column=col, value=val), numeric=False)
        for col, name in enumerate(line_names, 4):
            _apply_value(ws.cell(row=row, column=col, value=_lookup_line(lines, name)), numeric=True)
        row += 1

    for col, width in (("A", 14), ("B", 28), ("C", 14)):
        ws.column_dimensions[col].width = width
    for col in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "D3"


def _build_indicators(ws, results: list[TaskResult]) -> None:
    ws.title = "指标表"
    indicators = [
        ("摊薄每股收益(元)", lambda inc, bs, cf: _value(inc, "基本每股收益", "BASIC_EPS", "EarningsPerShareBasic")),
        ("加权每股收益(元)", lambda inc, bs, cf: _value(inc, "稀释每股收益", "EarningsPerShareDiluted")),
        ("每股收益_调整后(元)", lambda inc, bs, cf: _value(inc, "基本每股收益", "BASIC_EPS", "EarningsPerShareBasic")),
        ("主营业务收入增长率(%)", lambda inc, bs, cf: _value(inc, "营业收入同比", "营业总收入同比")),
        ("净资产收益率(%)", lambda inc, bs, cf: _scale_percent(_safe_div(_value(inc, "净利润", "PARENT_NETPROFIT", "NETPROFIT", "NetIncomeLoss"), _value(bs, "股东权益合计", "所有者权益(或股东权益)合计", "TOTAL_EQUITY", "StockholdersEquity")))),
        ("销售毛利率(%)", lambda inc, bs, cf: _scale_percent(_gross_margin(inc))),
        ("应收账款周转率(次)", lambda inc, bs, cf: _safe_div(_value(inc, "营业收入", "TOTAL_OPERATE_INCOME", "OPERATE_INCOME", "Revenues"), _value(bs, "应收账款", "ACCOUNTS_RECE", "AccountsReceivableNetCurrent"))),
        ("存货周转率(次)", lambda inc, bs, cf: _safe_div(_value(inc, "营业成本", "OPERATE_COST", "CostOfRevenue", "CostOfGoodsAndServicesSold"), _value(bs, "存货", "INVENTORY", "InventoryNet"))),
        ("总资产周转率(次)", lambda inc, bs, cf: _safe_div(_value(inc, "营业收入", "TOTAL_OPERATE_INCOME", "OPERATE_INCOME", "Revenues"), _value(bs, "资产总计", "TOTAL_ASSETS", "Assets"))),
        ("流动比率", lambda inc, bs, cf: _safe_div(_value(bs, "流动资产合计", "TOTAL_CURRENT_ASSETS", "AssetsCurrent"), _value(bs, "流动负债合计", "TOTAL_CURRENT_LIAB", "LiabilitiesCurrent"))),
        ("速动比率", lambda inc, bs, cf: _safe_div((_value(bs, "流动资产合计", "TOTAL_CURRENT_ASSETS", "AssetsCurrent") or 0) - (_value(bs, "存货", "INVENTORY", "InventoryNet") or 0), _value(bs, "流动负债合计", "TOTAL_CURRENT_LIAB", "LiabilitiesCurrent"))),
        ("现金比率(%)", lambda inc, bs, cf: _scale_percent(_safe_div(_value(bs, "货币资金", "MONETARYFUNDS", "CashAndCashEquivalentsAtCarryingValue", "Cash"), _value(bs, "流动负债合计", "TOTAL_CURRENT_LIAB", "LiabilitiesCurrent")))),
        ("股东权益比率(%)", lambda inc, bs, cf: _scale_percent(_safe_div(_value(bs, "股东权益合计", "所有者权益(或股东权益)合计", "TOTAL_EQUITY", "StockholdersEquity"), _value(bs, "资产总计", "TOTAL_ASSETS", "Assets")))),
        ("长期负债比率(%)", lambda inc, bs, cf: _scale_percent(_safe_div(_value(bs, "非流动负债合计", "LiabilitiesNoncurrent"), _value(bs, "资产总计", "TOTAL_ASSETS", "Assets")))),
        ("资产负债率(%)", lambda inc, bs, cf: _scale_percent(_safe_div(_value(bs, "负债合计", "TOTAL_LIABILITIES", "Liabilities"), _value(bs, "资产总计", "TOTAL_ASSETS", "Assets")))),
        ("总资产(元)", lambda inc, bs, cf: _value(bs, "资产总计", "TOTAL_ASSETS", "Assets")),
        ("经营现金净流量对销售收入比率(%)", lambda inc, bs, cf: _scale_percent(_safe_div(_value(cf, "经营活动现金流量净额", "NETCASH_OPERATE", "NetCashProvidedByUsedInOperatingActivities"), _value(inc, "营业收入", "TOTAL_OPERATE_INCOME", "OPERATE_INCOME", "Revenues")))),
        ("资产的经营现金流量回报率(%)", lambda inc, bs, cf: _scale_percent(_safe_div(_value(cf, "经营活动现金流量净额", "NETCASH_OPERATE", "NetCashProvidedByUsedInOperatingActivities"), _value(bs, "资产总计", "TOTAL_ASSETS", "Assets")))),
        ("经营现金净流量与净利润的比率(%)", lambda inc, bs, cf: _scale_percent(_safe_div(_value(cf, "经营活动现金流量净额", "NETCASH_OPERATE", "NetCashProvidedByUsedInOperatingActivities"), _value(inc, "净利润", "PARENT_NETPROFIT", "NETPROFIT", "NetIncomeLoss")))),
        ("经营现金净流量对负债比率(%)", lambda inc, bs, cf: _scale_percent(_safe_div(_value(cf, "经营活动现金流量净额", "NETCASH_OPERATE", "NetCashProvidedByUsedInOperatingActivities"), _value(bs, "负债合计", "TOTAL_LIABILITIES", "Liabilities")))),
        ("营业收入(亿)", lambda inc, bs, cf: _scale_billion(_value(inc, "营业收入", "TOTAL_OPERATE_INCOME", "OPERATE_INCOME", "Revenues"))),
        ("净利润(亿)", lambda inc, bs, cf: _scale_billion(_value(inc, "净利润", "PARENT_NETPROFIT", "NETPROFIT", "NetIncomeLoss"))),
        ("资产总额(亿)", lambda inc, bs, cf: _scale_billion(_value(bs, "资产总计", "TOTAL_ASSETS", "Assets"))),
        ("负债合计(亿)", lambda inc, bs, cf: _scale_billion(_value(bs, "负债合计", "TOTAL_LIABILITIES", "Liabilities"))),
        ("ROE", lambda inc, bs, cf: _safe_div(_value(inc, "净利润", "PARENT_NETPROFIT", "NETPROFIT", "NetIncomeLoss"), _value(bs, "股东权益合计", "所有者权益(或股东权益)合计", "TOTAL_EQUITY", "StockholdersEquity"))),
        ("毛利率", lambda inc, bs, cf: _gross_margin(inc)),
        ("资产负债率", lambda inc, bs, cf: _safe_div(_value(bs, "负债合计", "TOTAL_LIABILITIES", "Liabilities"), _value(bs, "资产总计", "TOTAL_ASSETS", "Assets"))),
        ("流动比率", lambda inc, bs, cf: _safe_div(_value(bs, "流动资产合计", "TOTAL_CURRENT_ASSETS", "AssetsCurrent"), _value(bs, "流动负债合计", "TOTAL_CURRENT_LIAB", "LiabilitiesCurrent"))),
        ("速动比率", lambda inc, bs, cf: _safe_div((_value(bs, "流动资产合计", "TOTAL_CURRENT_ASSETS", "AssetsCurrent") or 0) - (_value(bs, "存货", "INVENTORY", "InventoryNet") or 0), _value(bs, "流动负债合计", "TOTAL_CURRENT_LIAB", "LiabilitiesCurrent"))),
        ("现金比率", lambda inc, bs, cf: _safe_div(_value(bs, "货币资金", "MONETARYFUNDS", "CashAndCashEquivalentsAtCarryingValue", "Cash"), _value(bs, "流动负债合计", "TOTAL_CURRENT_LIAB", "LiabilitiesCurrent"))),
        ("应收账款周转率", lambda inc, bs, cf: _safe_div(_value(inc, "营业收入", "TOTAL_OPERATE_INCOME", "OPERATE_INCOME", "Revenues"), _value(bs, "应收账款", "ACCOUNTS_RECE", "AccountsReceivableNetCurrent"))),
        ("存货周转率", lambda inc, bs, cf: _safe_div(_value(inc, "营业成本", "OPERATE_COST", "CostOfRevenue", "CostOfGoodsAndServicesSold"), _value(bs, "存货", "INVENTORY", "InventoryNet"))),
        ("总资产周转率", lambda inc, bs, cf: _safe_div(_value(inc, "营业收入", "TOTAL_OPERATE_INCOME", "OPERATE_INCOME", "Revenues"), _value(bs, "资产总计", "TOTAL_ASSETS", "Assets"))),
        ("经营现金流/净利润", lambda inc, bs, cf: _safe_div(_value(cf, "经营活动现金流量净额", "NETCASH_OPERATE", "NetCashProvidedByUsedInOperatingActivities"), _value(inc, "净利润", "PARENT_NETPROFIT", "NETPROFIT", "NetIncomeLoss"))),
    ]
    indicator_map = {name: fn for name, fn in indicators}
    schema_names = _schema_columns("指标表")
    indicator_names = schema_names + [name for name, _ in indicators if name not in schema_names]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    _apply_section(ws.cell(row=1, column=1, value="基础信息"))
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=3 + len(indicator_names))
    _apply_section(ws.cell(row=1, column=4, value="财务指标"))
    headers = ["股票代码", "名称", "报告日期"] + indicator_names
    for col, h in enumerate(headers, 1):
        _apply_header(ws.cell(row=2, column=col, value=h))

    by_is: dict[tuple, dict] = {}
    by_bs: dict[tuple, dict] = {}
    by_cf: dict[tuple, dict] = {}
    for r in results:
        key = (r.ticker.exchange.value, r.ticker.code, r.period.year, r.period.type.value)
        for s in r.statements:
            if s.statement_type is StatementType.INCOME:
                by_is[key] = s.lines
            elif s.statement_type is StatementType.BALANCE_SHEET:
                by_bs[key] = s.lines
            elif s.statement_type is StatementType.CASH_FLOW:
                by_cf[key] = s.lines

    row = 3
    for r in results:
        key = (r.ticker.exchange.value, r.ticker.code, r.period.year, r.period.type.value)
        inc, bs, cf = by_is.get(key, {}), by_bs.get(key, {}), by_cf.get(key, {})
        if not inc and not bs and not cf:
            continue
        for col, val in enumerate([r.ticker.code, r.ticker.name or "", _period_date(r.period)], 1):
            _apply_value(ws.cell(row=row, column=col, value=val), numeric=False)
        for col, name in enumerate(indicator_names, 4):
            val = None
            fn = indicator_map.get(name)
            try:
                val = fn(inc, bs, cf) if fn else (_lookup_line(inc, name) or _lookup_line(bs, name) or _lookup_line(cf, name))
            except Exception:
                val = None
            cell = ws.cell(row=row, column=col, value=val)
            _apply_value(cell, numeric=True)
            if name in _PERCENT_INDICATORS:
                cell.number_format = "0.00%"
        row += 1

    for col, width in (("A", 14), ("B", 28), ("C", 14)):
        ws.column_dimensions[col].width = width
    for col in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "D3"


def _build_reports_index(ws, results: list[TaskResult]) -> None:
    ws.title = "报告下载汇总"
    headers = ["交易所", "代码", "公司名称", "期间", "报告类型", "标题", "文件大小(KB)", "本地路径", "源URL"]
    for i, h in enumerate(headers, 1):
        _apply_header(ws.cell(row=1, column=i, value=h))

    row = 2
    for r in results:
        for f in r.reports:
            values = [
                r.ticker.exchange.display_name,
                r.ticker.code,
                r.ticker.name or "",
                r.period.label(),
                f.kind,
                f.title or "",
                round((f.file_size_bytes or 0) / 1024, 1),
                f.local_path,
                f.source_url,
            ]
            for col, val in enumerate(values, 1):
                _apply_value(ws.cell(row=row, column=col, value=val), numeric=isinstance(val, (int, float)))
            row += 1

    widths = [8, 12, 24, 14, 14, 36, 12, 50, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


@dataclass
class WriteResult:
    path: Path
    n_columns: int
    n_reports: int


def write_workbook(job: Job, out_path: Path) -> WriteResult:
    results = list(job.results)
    successful = [r for r in results if r.status is TaskStatus.DONE]

    wb = Workbook()
    wb.remove(wb.active)
    _build_sample_pool(wb.create_sheet(), results)
    _build_company_info(wb.create_sheet(), results)
    _build_indicators(wb.create_sheet(), successful)
    _build_statement_sheet(wb.create_sheet(), "资产负债表", StatementType.BALANCE_SHEET, successful)
    _build_statement_sheet(wb.create_sheet(), "利润表", StatementType.INCOME, successful)
    _build_statement_sheet(wb.create_sheet(), "现金流量表", StatementType.CASH_FLOW, successful)
    _build_reports_index(wb.create_sheet(), results)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    n_cols = len(_columns_in_order(successful))
    n_reports = sum(len(r.reports) for r in results)
    logger.info(f"workbook written: {out_path} (cols={n_cols}, reports={n_reports})")
    return WriteResult(path=out_path, n_columns=n_cols, n_reports=n_reports)
