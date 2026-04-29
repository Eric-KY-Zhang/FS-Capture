from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.core.job import Job, TaskResult, TaskStatus
from app.core.models import (
    Exchange,
    FinancialStatement,
    Period,
    PeriodType,
    StatementType,
    Ticker,
)
from app.core.output_paths import report_output_path
from app.exporters.excel_writer import write_workbook


class OutputLayoutTests(unittest.TestCase):
    def test_report_path_is_flat_under_output_root(self) -> None:
        ticker = Ticker(exchange=Exchange.A_SHARE, code="600519")
        period = Period(year=2024, type=PeriodType.ANNUAL)

        path = report_output_path(Path("output"), ticker, period, "annual_report", ".pdf")

        self.assertEqual(path.parent, Path("output"))
        self.assertEqual(path.name, "A_600519_2024_annual_annual_report.pdf")

    def test_workbook_uses_visible_wide_sheets_without_ruihua_tab(self) -> None:
        ticker = Ticker(exchange=Exchange.A_SHARE, code="600519", name="贵州茅台")
        period = Period(year=2024, type=PeriodType.ANNUAL)
        statements = [
            FinancialStatement(
                ticker=ticker,
                period=period,
                statement_type=StatementType.BALANCE_SHEET,
                lines={"TOTAL_ASSETS": 1000.0, "TOTAL_LIABILITIES": 400.0, "TOTAL_EQUITY": 600.0},
            ),
            FinancialStatement(
                ticker=ticker,
                period=period,
                statement_type=StatementType.INCOME,
                lines={"TOTAL_OPERATE_INCOME": 200.0, "NETPROFIT": 50.0, "OPERATE_COST": 80.0},
            ),
            FinancialStatement(
                ticker=ticker,
                period=period,
                statement_type=StatementType.CASH_FLOW,
                lines={"NETCASH_OPERATE": 45.0},
            ),
        ]
        job = Job(tickers=[ticker], periods=[period], output_dir="")
        job.results.append(TaskResult(ticker=ticker, period=period, status=TaskStatus.DONE, statements=statements))

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "workbook.xlsx"
            write_workbook(job, out)
            wb = load_workbook(out, data_only=False)

        self.assertNotIn("瑞华底稿", wb.sheetnames)
        self.assertIn("指标表", wb.sheetnames)
        self.assertIn("资产负债表", wb.sheetnames)
        self.assertEqual(wb["指标表"]["A2"].value, "股票代码")
        self.assertEqual(wb["指标表"]["D2"].value, "摊薄每股收益(元)")
        self.assertEqual(wb["资产负债表"]["D2"].value, "货币资金")
        self.assertEqual(wb["指标表"]["A3"].value, "600519")


if __name__ == "__main__":
    unittest.main()
