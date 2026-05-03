"""
上市公司财务数据查询 — 工作簿模板生成器

生成空的 xlsm 模板,含使用说明、样本池、A股 4 表、美股 4 表、港股 4 表的结构 + 列宽 + 表头容器 + 冻结窗格。
不含任何 VBA 代码 — 后续由 install_modules.py 注入 modules/*.bas。

用法:
    cd "E:\\Claude+CODEX Project\\FS Capture\\VBA Captor"
    py tools/build_template.py

输出:
    上市公司财务数据查询.xlsx  (在工作目录根)

注意:
    输出 .xlsx 而非 .xlsm — openpyxl 给 fresh 工作簿写 .xlsm 时不带 vbaProject,
    Excel 会拒绝打开。后续由 tools/install_modules.py 用 Excel COM 转成 .xlsm
    并注入 VBA 模块。
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_PATH = Path(__file__).resolve().parent.parent / "上市公司财务数据查询.xlsx"

DARK_BLUE = "FF4472C4"
LIGHT_GRAY = "FFD9D9D9"
WHITE = "FFFFFFFF"

THIN = Side(border_style="thin", color="FF808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
SUB_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FF000000")
DATA_FONT = Font(name="微软雅黑", size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_intro(ws):
    ws.column_dimensions["A"].width = 100
    ws["A1"] = "上市公司财务数据查询"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True)

    lines = [
        "",
        "【用途】把上市公司财务数据抓成同业对标宽表,方便横向比较。",
        "【当前支持】A股、美股、港股。",
        "【后续规划】韩股。",
        "【作者】Eric Zhang",
        "【联系邮箱】214978902@qq.com",
        "",
        "【使用步骤】",
        "1. 在『样本池』Sheet 第 8 行起录入公司:",
        "     A 列代码, B 列简称, C 列市场 (A / US / HK / KR; 留空可自动判断 A/US/HK, KR 请手填)。",
        "2. A2 填年份 (如 2025), 留空=取最新可用期间。",
        "3. A4 选择季度: 全部 / Q1 / Q2 / Q3 / Q4。",
        "4. B5 可填写雪球 xq_a_token cookie; POM、HTT 等 EDGAR 不完整的中概/20-F 公司会自动走雪球 fallback, 港股也使用该 cookie。",
        "5. 点样本池右上角按钮抓数:",
        "     【一键全抓】 — 顺序更新 A股、美股、港股 12 张表, 最后弹一次汇总",
        "     【更新A股资产负债表 / 利润表 / 现金流量表 / 指标表】",
        "     【更新美股资产负债表 / 利润表 / 现金流量表 / 指标表】",
        "     【更新港股资产负债表 / 利润表 / 现金流量表 / 指标表】",
        "",
        "【宽表格式】",
        "  R1: 公司名(代码), 跨该公司所有报告期合并",
        "  R2: 报告期, 降序排列",
        "  A/B列: 大类或指标类型、指标名称; 指标表额外有 C列英文指标名",
        "  A股: 报告期跨公司取并集对齐; 美股/港股: 每家公司只展开自己有数据的报告期",
        "  港股: 单位为百万(各家公司报告币种, 见 港股_抓取诊断 Unit 列), 不做汇率换算",
        "",
        "【数据源】",
        "  A股: 新浪财经",
        "  美股: SEC EDGAR companyfacts; 中概/20-F fallback 到雪球",
        "  港股: 雪球 HK API",
        "",
        "【限制】",
        "  - 雪球 cookie 过期时需重新复制 xq_a_token 到 B5",
        "  - 韩股目前为规划市场, 尚未接入抓数",
        "",
        "【来源说明】基于林铖 V2.2 重写并扩展。",
    ]
    for i, text in enumerate(lines, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = LEFT
        cell.font = Font(name="微软雅黑", size=11)


def build_sample_pool(ws):
    """
    样本池布局 (Phase 4b-3 简化版):
      Row 1: A1=年份标签
      Row 2: A2=年份值 (留空=取最新季度)
      Row 3: A3=季度标签
      Row 4: A4=季度值 (下拉: 全部/Q1/Q2/Q3/Q4)
      Row 7: 数据表头 (A=股票代码 / B=股票简称 / C=市场)
      Row 8+: 股票数据
      D 列起: 按钮区 (install_modules.py 装的圆角按钮)

    URL 不再存 sheet — 各抓数模块内部按代码 + A2 年份自拼新浪 URL
    """
    sub_header_fill = PatternFill("solid", fgColor="FFB4C7E7")    # 浅蓝
    main_header_fill = PatternFill("solid", fgColor=DARK_BLUE)

    # ---- 列宽 (D=spacer 让市场列和按钮列之间有视觉间隔) ----
    col_widths = {"A": 13, "B": 16, "C": 10, "D": 3, "E": 24, "F": 4}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ---- Row 1 / Row 3: 配置区标签 ----
    for addr, txt in [
        ("A1", "年份 (留空=取最新)"),
        ("A3", "季度 (Q1/Q2/Q3/Q4 或 全部)"),
    ]:
        ws[addr] = txt
        ws[addr].font = Font(name="微软雅黑", size=10, bold=True)
        ws[addr].fill = sub_header_fill
        ws[addr].alignment = CENTER
        ws[addr].border = BORDER

    # ---- A2: 年份默认 2025 ----
    ws["A2"] = 2025
    ws["A2"].font = Font(name="微软雅黑", size=11, bold=True)
    ws["A2"].fill = PatternFill("solid", fgColor="FFFFE699")
    ws["A2"].alignment = CENTER
    ws["A2"].border = BORDER

    # ---- A4: 季度默认 全部, 下拉验证 ----
    ws["A4"] = "全部"
    ws["A4"].font = Font(name="微软雅黑", size=11, bold=True)
    ws["A4"].fill = PatternFill("solid", fgColor="FFFFE699")
    ws["A4"].alignment = CENTER
    ws["A4"].border = BORDER
    dv = DataValidation(type="list", formula1='"全部,Q1,Q2,Q3,Q4"', allow_blank=False)
    dv.add("A4")
    ws.add_data_validation(dv)

    # ---- Row 7: 数据表头 ----
    data_headers = [("A", "股票代码"), ("B", "股票简称"), ("C", "市场")]
    for col, name in data_headers:
        cell = ws[f"{col}7"]
        cell.value = name
        cell.font = HEADER_FONT
        cell.fill = main_header_fill
        cell.alignment = CENTER
        cell.border = BORDER

    ws.row_dimensions[7].height = 22
    ws.freeze_panes = "A8"

    # ---- Row 8: 示例数据 (300866 安克创新) ----
    ws.cell(row=8, column=1, value=300866).alignment = CENTER
    ws.cell(row=8, column=2, value="安克创新").alignment = CENTER
    # C8 自动检测公式 (install_modules.py 也会复读 / 覆盖 H8:H1000)
    formula = '=IF(A8="","",IF(ISNUMBER(--A8),IF(LEN(A8)=5,"HK","A"),"US"))'
    ws.cell(row=8, column=3, value=formula).alignment = CENTER


def build_wide_table(ws):
    """
    宽表 Sheet 的初始结构:
      A1=大类, B1=指标名称, R2 留空给 VBA 填日期, R1 C+ 给 VBA 填公司名
      列宽 A=30, B=40, C+=15.875 (但 C+ 只设默认列宽,VBA 写时会扩列)
      冻结 B3 (前 2 行表头 + 前 2 列锚定)
    """
    fill = PatternFill("solid", fgColor=DARK_BLUE)

    ws["A1"] = "大类"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = fill
    ws["A1"].alignment = CENTER
    ws["A1"].border = BORDER

    ws["B1"] = "指标名称"
    ws["B1"].font = HEADER_FONT
    ws["B1"].fill = fill
    ws["B1"].alignment = CENTER
    ws["B1"].border = BORDER

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15.875
    ws.sheet_format.defaultColWidth = 15.875

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "C3"


def build_corp_info(ws):
    headers = [
        ("A", "股票代码", 12),
        ("B", "股票简称", 14),
        ("C", "上市日期", 14),
        ("D", "所属行业", 24),
        ("E", "主营业务", 80),
    ]
    fill = PatternFill("solid", fgColor=DARK_BLUE)
    for col, name, width in headers:
        cell = ws[f"{col}1"]
        cell.value = name
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def build_diagnostic_sheet(ws, market_label="美股"):
    """
    Phase 4b-14a/4c: 抓取诊断 sheet 模板
      Row 1: 大标题 (合并 A1:J1, 深蓝白字)
      Row 2: 10 列表头 — 公司/报表/输出指标/状态/数据源/Taxonomy/命中字段/Unit/Score/匹配方式+备注
      Row 3+: 由 VBA 写入(每次跑数后刷新)
      冻结 Row 2; 列宽 + 表头颜色与 install_modules.py._make_diagnostic_sheet
        和 模块_工具函数.bas.EnsureDiagnosticSheet() 三处保持一致
    """
    title_fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_font = Font(name="微软雅黑", size=12, bold=True, color=WHITE)

    # Row 1: 标题, 合并 A1:J1
    ws["A1"] = f"{market_label}抓取诊断 (每次跑数后自动刷新)"
    ws.merge_cells("A1:J1")
    for cell_addr in ["A1"]:
        c = ws[cell_addr]
        c.font = title_font
        c.fill = title_fill
        c.alignment = CENTER
        c.border = BORDER

    # Row 2: 10 列表头
    headers = ["公司", "报表", "输出指标", "状态", "数据源",
               "Taxonomy", "命中字段", "Unit", "Score", "匹配方式+备注"]
    header_font = Font(name="微软雅黑", size=10, bold=True, color=WHITE)
    for j, txt in enumerate(headers, start=1):
        col_letter = get_column_letter(j)
        c = ws[f"{col_letter}2"]
        c.value = txt
        c.font = header_font
        c.fill = title_fill
        c.alignment = CENTER
        c.border = BORDER

    widths = [14, 16, 30, 18, 18, 14, 42, 14, 10, 58]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20

    # 冻结 Row 2 (滚动时表头常驻)
    ws.freeze_panes = "A3"


def main():
    wb = Workbook()
    # Drop the default sheet
    wb.remove(wb.active)

    # 1. 使用说明
    ws_intro = wb.create_sheet("使用说明")
    build_intro(ws_intro)

    # 2. 样本池
    ws_pool = wb.create_sheet("样本池")
    build_sample_pool(ws_pool)

    # ---- A 股 4 张表 (Phase 4b-4: 加 A股_ 前缀, 跟美股对称) ----
    for name in ["A股_资产负债表", "A股_利润表", "A股_现金流量表", "A股_指标表"]:
        ws = wb.create_sheet(name)
        build_wide_table(ws)

    # ---- Phase 4b: 美股 4 张表 (单位: 百万美元 / EPS 美元/股 / 加权股数 百万股) ----
    for name in ["美股_资产负债表", "美股_利润表", "美股_现金流量表", "美股_指标表"]:
        ws_us = wb.create_sheet(name)
        build_wide_table(ws_us)

    # ---- Phase 4b-14a: 美股_抓取诊断 sheet (放最后) ----
    ws_diag = wb.create_sheet("美股_抓取诊断")
    build_diagnostic_sheet(ws_diag, "美股")

    # ---- Phase 4c: 港股 4 张表 + 港股_抓取诊断 sheet ----
    for name in ["港股_资产负债表", "港股_利润表", "港股_现金流量表", "港股_指标表"]:
        ws_hk = wb.create_sheet(name)
        build_wide_table(ws_hk)

    ws_diag_hk = wb.create_sheet("港股_抓取诊断")
    build_diagnostic_sheet(ws_diag_hk, "港股")

    # 默认打开时显示样本池
    wb.active = wb.sheetnames.index("样本池")

    wb.save(OUT_PATH)
    print(f"Generated: {OUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
